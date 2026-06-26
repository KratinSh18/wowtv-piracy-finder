"""
Extract the WOW TV show catalog from the Live Show Tracker .xlsx into shows.csv.

Pulls titles from every useful tab and dedupes them:
  - "Live Tracker"        -> Show Name (+ Genre / Tropes for context)
  - "base data"           -> show_name AND ip_title (the original English title a
                             dubbed show is based on -- pirates use either)
  - "base data_ML_1706"   -> Show Title

Re-run this whenever the sheet changes:
    python extract_catalog.py "C:\\path\\to\\WowTV Show Live Tracker.xlsx"
Output: shows.csv  (a 'title' column the tool reads, plus language/show_id/etc.)
"""
from __future__ import annotations

import csv
import re
import sys

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

DEFAULT_XLSX = r"C:\Users\admin\Downloads\Copy of WowTV Show Live Tracker.xlsx"
BAD = {"", "#ref!", "#n/a", "null", "none", "nan", "show name", "show title",
       "ip_title", "show_name"}


def clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def find_header(ws, wanted, max_scan=8):
    """Return (header_row_index, {colname_lower: col_index}) for the first row
    that contains any of the wanted header names."""
    rows = list(ws.iter_rows(values_only=True))
    wl = {w.lower() for w in wanted}
    for i, row in enumerate(rows[:max_scan]):
        cells = [clean(c).lower() for c in row]
        if wl & set(cells):
            idx = {c: j for j, c in enumerate(cells) if c}
            return i, idx, rows
    return None, {}, rows


def add(catalog, title, **meta):
    t = clean(title)
    if not t or t.lower() in BAD or t.startswith("%sql") or len(t) < 2:
        return
    key = re.sub(r"[^a-z0-9]+", " ", t.lower()).strip()
    if not key:
        return
    if key not in catalog:
        catalog[key] = {"title": t, "language": "", "show_id": "",
                        "genre": "", "tropes": "", "sources": set()}
    rec = catalog[key]
    for k in ("language", "show_id", "genre", "tropes"):
        if not rec[k] and meta.get(k):
            rec[k] = clean(meta[k])
    if meta.get("source"):
        rec["sources"].add(meta["source"])


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    catalog = {}

    # ---- Live Tracker ----
    if "Live Tracker" in wb.sheetnames:
        ws = wb["Live Tracker"]
        hi, idx, rows = find_header(ws, ["Show Name", "Show ID"])
        if hi is not None:
            c_name = idx.get("show name")
            c_lang = idx.get("language")
            c_id = idx.get("show id")
            c_gen = idx.get("genre")
            c_t1, c_t2 = idx.get("tropes 1"), idx.get("tropes 2")
            for row in rows[hi + 1:]:
                if c_name is None or c_name >= len(row):
                    continue
                tropes = " / ".join(clean(row[c]) for c in (c_t1, c_t2)
                                    if c is not None and c < len(row) and clean(row[c]))
                add(catalog, row[c_name],
                    language=row[c_lang] if c_lang is not None and c_lang < len(row) else "",
                    show_id=row[c_id] if c_id is not None and c_id < len(row) else "",
                    genre=row[c_gen] if c_gen is not None and c_gen < len(row) else "",
                    tropes=tropes, source="live_tracker")

    # ---- base data (show_name + ip_title) ----
    if "base data" in wb.sheetnames:
        ws = wb["base data"]
        hi, idx, rows = find_header(ws, ["show_name", "ip_title"])
        if hi is not None:
            c_name = idx.get("show_name")
            c_ip = idx.get("ip_title")
            c_lang = idx.get("language")
            c_id = idx.get("show_id")
            for row in rows[hi + 1:]:
                lang = row[c_lang] if c_lang is not None and c_lang < len(row) else ""
                sid = row[c_id] if c_id is not None and c_id < len(row) else ""
                if c_name is not None and c_name < len(row):
                    add(catalog, row[c_name], language=lang, show_id=sid, source="base_data")
                if c_ip is not None and c_ip < len(row):
                    add(catalog, row[c_ip], source="ip_title")

    # ---- base data_ML_1706 (Show Title) ----
    if "base data_ML_1706" in wb.sheetnames:
        ws = wb["base data_ML_1706"]
        hi, idx, rows = find_header(ws, ["Show Title"])
        if hi is not None:
            c_name = idx.get("show title")
            c_id = idx.get("showid")
            for row in rows[hi + 1:]:
                if c_name is not None and c_name < len(row):
                    add(catalog, row[c_name],
                        show_id=row[c_id] if c_id is not None and c_id < len(row) else "",
                        source="ml_base")

    rows_out = sorted(catalog.values(), key=lambda r: r["title"].lower())
    with open("shows.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["title", "language", "show_id", "genre", "tropes", "sources"])
        for r in rows_out:
            w.writerow([r["title"], r["language"], r["show_id"], r["genre"],
                        r["tropes"], ",".join(sorted(r["sources"]))])
    print(f"Extracted {len(rows_out)} unique titles -> shows.csv")
    print("First 15:")
    for r in rows_out[:15]:
        print(f"  - {r['title']}  ({r['language'] or '?'}; {','.join(sorted(r['sources']))})")


if __name__ == "__main__":
    main()
